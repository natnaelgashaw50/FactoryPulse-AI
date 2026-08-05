import os
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from backend.extensions import db
from backend.models import Machine, Alert, Prediction, Report

bp = Blueprint("reports", __name__, url_prefix="/api/reports")

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "generated_reports")
os.makedirs(REPORTS_DIR, exist_ok=True)


def _period_start(period: str) -> datetime:
    now = datetime.utcnow()
    return {"daily": now - timedelta(days=1), "weekly": now - timedelta(weeks=1),
            "monthly": now - timedelta(days=30)}.get(period, now - timedelta(days=1))


def _gather(period: str):
    since = _period_start(period)
    machines = Machine.query.all()
    alerts = Alert.query.filter(Alert.created_at >= since).all()
    predictions = Prediction.query.filter(Prediction.created_at >= since).all()
    return machines, alerts, predictions


@bp.post("/generate")
@jwt_required()
def generate_report():
    data = request.get_json() or {}
    period = data.get("period", "daily")
    fmt = data.get("format", "pdf")
    machines, alerts, predictions = _gather(period)

    filename = f"ishfp_{period}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{fmt}"
    path = os.path.join(REPORTS_DIR, filename)

    if fmt == "pdf":
        _build_pdf(path, period, machines, alerts, predictions)
    else:
        _build_xlsx(path, period, machines, alerts, predictions)

    report = Report(period=period, format=fmt, file_path=path)
    db.session.add(report)
    db.session.commit()
    return jsonify(report.to_dict()), 201


@bp.get("/<int:report_id>/download")
@jwt_required()
def download_report(report_id):
    report = Report.query.get_or_404(report_id)
    return send_file(report.file_path, as_attachment=True)


@bp.get("")
@jwt_required()
def list_reports():
    return jsonify([r.to_dict() for r in Report.query.order_by(Report.generated_at.desc()).all()])


def _build_pdf(path, period, machines, alerts, predictions):
    doc = SimpleDocTemplate(path, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [Paragraph(f"ISHFP {period.capitalize()} Report", styles["Title"]), Spacer(1, 12)]

    story.append(Paragraph("Machine Status", styles["Heading2"]))
    m_data = [["Name", "Zone", "Status", "Health Score"]] + [
        [m.name, m.zone, m.status, f"{m.health_score:.0f}"] for m in machines
    ]
    story.append(Table(m_data, style=TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ])))
    story.append(Spacer(1, 16))

    story.append(Paragraph(f"Alerts ({len(alerts)})", styles["Heading2"]))
    a_data = [["Title", "Severity", "Status", "Created"]] + [
        [a.title, a.severity, a.status, a.created_at.strftime("%Y-%m-%d %H:%M")] for a in alerts
    ] or [["No alerts in this period", "", "", ""]]
    story.append(Table(a_data, style=TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ])))

    doc.build(story)


def _build_xlsx(path, period, machines, alerts, predictions):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Machines"
    ws1.append(["Name", "Zone", "Status", "Health Score", "Running Hours"])
    for m in machines:
        ws1.append([m.name, m.zone, m.status, m.health_score, m.running_hours])

    ws2 = wb.create_sheet("Alerts")
    ws2.append(["Title", "Severity", "Status", "Machine ID", "Created At"])
    for a in alerts:
        ws2.append([a.title, a.severity, a.status, a.machine_id, a.created_at.isoformat()])

    ws3 = wb.create_sheet("Predictions")
    ws3.append(["Machine ID", "Failure Probability", "RUL Days", "Risk Level", "Created At"])
    for p in predictions:
        ws3.append([p.machine_id, p.failure_probability, p.remaining_useful_life_days, p.risk_level, p.created_at.isoformat()])

    wb.save(path)
